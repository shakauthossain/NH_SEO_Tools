"""
RankMath SEO Analyzer Scraper
------------------------------
Reads URLs from an Excel file, analyzes each one using
rankmath.com/tools/seo-analyzer/, and writes results back to Excel.

Requirements:
    pip install playwright openpyxl
    playwright install chromium
"""

import time
import openpyxl
from playwright.sync_api import sync_playwright


# ─── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE  = "leads.xlsx"       # Your input Excel file
OUTPUT_FILE = "leads_seo.xlsx"   # Output file with SEO results
URL_COLUMN  = 1                  # Column index (1 = A) where URLs are stored
START_ROW   = 2                  # Row to start from (2 = skip header)
DELAY       = 8                  # Seconds to wait between each URL (avoid blocking)
HEADLESS    = True               # Set False to watch the browser in action
# ───────────────────────────────────────────────────────────────────────────────


def get_seo_data(page, url):
    """Submit a URL to RankMath analyzer and scrape the results."""
    try:
        print(f"  → Analyzing: {url}")

        # Go to the tool page
        page.goto("https://rankmath.com/tools/seo-analyzer/", timeout=30000)
        page.wait_for_load_state("networkidle")

        # Find the input field and type the URL
        input_box = page.locator("input#url")
        input_box.wait_for(timeout=10000)
        input_box.fill("")
        input_box.fill(url)

        # Click the Analyze button
        page.locator("button#analyze").click()

        # Wait for results container to appear
        page.wait_for_selector("div.container.analysis-result.clear", timeout=60000)
        page.wait_for_load_state("networkidle")
        time.sleep(2)  # Extra buffer for all numbers to render

        # ── Scrape SEO Score ──────────────────────────────────────────────────
        score_text = page.locator("div.rank-math-result-graphs .score-value").inner_text()
        # Score appears as "65/100" — extract just the number
        seo_score = score_text.split("/")[0].strip() if "/" in score_text else score_text.strip()

        # ── Scrape Passed / Warnings / Failed ────────────────────────────────
        graphs = page.locator("div.rank-math-result-graphs")
        all_text = graphs.inner_text()

        # Parse the values from text like "18/28\nPassed Tests\n4/28\nWarnings\n6/28\nFailed Tests"
        passed   = extract_fraction(all_text, "Passed Tests")
        warnings = extract_fraction(all_text, "Warnings")
        failed   = extract_fraction(all_text, "Failed Tests")

        print(f"     Score: {seo_score}/100 | Passed: {passed} | Warnings: {warnings} | Failed: {failed}")
        return seo_score, passed, warnings, failed

    except Exception as e:
        print(f"     ✗ Error: {e}")
        return "Error", "Error", "Error", "Error"


def extract_fraction(text, label):
    """Extract the X/Y fraction that appears before a label in the text block."""
    try:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        for i, line in enumerate(lines):
            if label in line and i > 0:
                return lines[i - 1]  # The fraction is the line before the label
    except Exception:
        pass
    return "N/A"


def main():
    # ── Load workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # ── Write headers for new columns ─────────────────────────────────────────
    header_row   = START_ROW - 1
    last_col     = ws.max_column + 1
    ws.cell(row=header_row, column=last_col,     value="SEO Score (/100)")
    ws.cell(row=header_row, column=last_col + 1, value="Passed Tests")
    ws.cell(row=header_row, column=last_col + 2, value="Warnings")
    ws.cell(row=header_row, column=last_col + 3, value="Failed Tests")

    # ── Collect URLs ──────────────────────────────────────────────────────────
    rows_to_process = []
    for row in range(START_ROW, ws.max_row + 1):
        url = ws.cell(row=row, column=URL_COLUMN).value
        if url:
            # Ensure URL has a scheme
            url = url.strip()
            if not url.startswith("http"):
                url = "https://" + url
            rows_to_process.append((row, url))

    print(f"\n📋 Found {len(rows_to_process)} URLs to analyze\n")

    # ── Launch Playwright ─────────────────────────────────────────────────────
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        for i, (row, url) in enumerate(rows_to_process, 1):
            print(f"[{i}/{len(rows_to_process)}]")

            score, passed, warnings, failed = get_seo_data(page, url)

            # Write results back to the row
            ws.cell(row=row, column=last_col,     value=score)
            ws.cell(row=row, column=last_col + 1, value=passed)
            ws.cell(row=row, column=last_col + 2, value=warnings)
            ws.cell(row=row, column=last_col + 3, value=failed)

            # Save after every URL so you don't lose progress
            wb.save(OUTPUT_FILE)

            # Delay to avoid getting rate-limited or blocked
            if i < len(rows_to_process):
                print(f"     ⏳ Waiting {DELAY}s before next URL...")
                time.sleep(DELAY)

        browser.close()

    print(f"\n✅ Done! Results saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
